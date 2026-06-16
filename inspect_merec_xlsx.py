import openpyxl
import os

def inspect():
    path = "C:\\laragon\\www\\SPKTUBES"
    file_path = os.path.join(path, "MEREC_OCRA_SPK_Lengkap_ori.xlsx")
    if not os.path.exists(file_path):
        print("File not found!")
        return
    wb = openpyxl.load_workbook(file_path, data_only=False)
    print("Sheets:", wb.sheetnames)
    
    # Check if there is a sheet related to MEREC weights
    for name in wb.sheetnames:
        if "merec" in name.lower() or "bobot" in name.lower() or "weight" in name.lower():
            sheet = wb[name]
            print(f"\n=== Sheet: {name} ===")
            print("Dimensions:", sheet.dimensions)
            # Print first 10 rows and columns
            for r in range(1, 15):
                row_vals = [sheet.cell(r, c).value for c in range(1, 10)]
                print(f"  Row {r}: {row_vals}")

if __name__ == "__main__":
    inspect()
